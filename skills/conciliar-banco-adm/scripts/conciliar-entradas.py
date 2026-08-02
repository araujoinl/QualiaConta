#!/usr/bin/env python3
"""
Conciliación de Entradas Bancarias — Blackbox SRL

Script standalone que descarga todo (Supabase + ADM Cloud), cruza por rondas
(incluyendo tarjetas de crédito), y genera Excel + JSON.

USO:
    python3 conciliar-entradas.py [YYYY-MM] [YYYY-MM]

    Ejemplo:
    python3 conciliar-entradas.py                    # mes actual
    python3 conciliar-entradas.py 2026-07            # julio 2026
    python3 conciliar-entradas.py 2026-06 2026-07    # rango jun-jul

REQUISITOS:
    - Variables de entorno: OPENBANKING_DSN, ADMCLOUD_COMPANY,
      ADMCLOUD_USER, ADMCLOUD_PASSWORD, ADMCLOUD_ROLE, ADMCLOUD_APPID
    - openpyxl instalado en algún venv accesible

IMPORTANTE: Este script se generó en una sesión interactiva y luego se
refinó. Para el algoritmo de cruce reutilizable, ver match-algorithm.py.
"""

import base64
import csv
import json
import os
import re
import sys
import unicodedata
import urllib.parse
import urllib.request
from datetime import datetime, date, timedelta

# ======================================================================
# CONFIGURACIÓN
# ======================================================================

CUENTAS_BLACKBOX = [
    "11121000000801",  # Ingresos DOP
    "11122010014964",  # Impuestos DOP
    "11122010023874",  # Operaciones DOP
    "21122020001404",  # Suplidores USD
    "21122020002181",  # Ganancias USD
]

CUENTA_NOMBRES = {
    "11121000000801": "Ingresos",
    "11122010014964": "Impuestos",
    "11122010023874": "Operaciones",
    "21122020001404": "Suplidores USD",
    "21122020002181": "Ganancias USD",
}

COMISION_TARJETA = 0.05395
PATRON_TARJETA = "servicios digita"

MONTO_EXACTO_DIFF = 0.50
MONTO_CERCANO_PCT = 0.005
MONTO_FUZZY_PCT   = 0.01
FECHA_MAX_EXACTO  = 10
FECHA_MAX_CERCANO = 7
FECHA_MAX_FUZZY   = 5

SUFFIXES = re.compile(
    r'\b(s\.?\s*a\.?\s*s?|s\.?\s*r\.?\s*l?|srl|sas|sa|s\.a\.?|'
    r'c\.?\s*por\s*a\.?|corporation|corp|inc|ltda|company|co)\b',
    re.IGNORECASE
)
BANK_WORDS = re.compile(
    r'\b(transferencia|recibida|de|env|dev|por|factura|vencidas|'
    r'credito|debito|ach|lbtr|fen|imp|comision|transf)\b',
    re.IGNORECASE
)


# ======================================================================
# UTILIDADES (mismas que match-algorithm.py)
# ======================================================================

def normalize_name(name):
    if not name or not name.strip():
        return ""
    name = unicodedata.normalize("NFD", name)
    name = name.encode("ascii", "ignore").decode("ascii")
    name = name.lower().strip()
    name = SUFFIXES.sub("", name)
    name = BANK_WORDS.sub("", name)
    name = re.sub(r"\d+", "", name)
    name = re.sub(r"[^a-z\s]", "", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def names_match(bank_desc, adm_name):
    bn = normalize_name(bank_desc)
    an = normalize_name(adm_name)
    if not bn or not an:
        return False
    if bn == an:
        return True
    if bn in an or an in bn:
        return True
    bw = set(w for w in bn.split() if len(w) > 2)
    aw = set(w for w in an.split() if len(w) > 2)
    if bw and aw and (bw & aw):
        return True
    return False


def is_credit_card(desc):
    return PATRON_TARJETA in desc.lower()


def date_diff_days(d1, d2):
    try:
        return abs(
            (datetime.strptime(d1, "%Y-%m-%d") - datetime.strptime(d2, "%Y-%m-%d")).days
        )
    except (ValueError, TypeError):
        return 999


# ======================================================================
# SUPABASE — Leer transacciones bancarias
# ======================================================================

def load_bank_txs(start_date, end_date):
    import subprocess
    in_list = ",".join("'" + c + "'" for c in CUENTAS_BLACKBOX)
    query = (
        f"SELECT fecha_posteo, cuenta_numero, descripcion, monto, balance, "
        f"nro_referencia, nombre_origen "
        f"FROM openbanking_transactions "
        f"WHERE cuenta_numero IN ({in_list}) "
        f"AND fecha_posteo >= '{start_date}' "
        f"AND fecha_posteo <= '{end_date}' "
        f"AND monto > 0 "
        f"ORDER BY fecha_posteo DESC;"
    )
    result = subprocess.run(
        ["psql", os.environ["OPENBANKING_DSN"], "--csv", "-c", query],
        capture_output=True, text=True, timeout=30,
    )
    if result.returncode != 0:
        print(f"Error leyendo Supabase: {result.stderr}", file=sys.stderr)
        sys.exit(1)
    txs = []
    for row in csv.DictReader(result.stdout.splitlines()):
        tx = {
            "fecha": row["fecha_posteo"],
            "cuenta": row["cuenta_numero"],
            "descripcion": row["descripcion"].strip(),
            "monto": float(row["monto"]),
            "balance": float(row["balance"]) if row["balance"] else None,
            "nro_referencia": row["nro_referencia"],
            "nombre_origen": row["nombre_origen"],
        }
        tx["is_credit_card"] = is_credit_card(tx["descripcion"])
        tx["monto_original"] = (
            tx["monto"] / (1 - COMISION_TARJETA)
            if tx["is_credit_card"]
            else tx["monto"]
        )
        txs.append(tx)
    return txs


# ======================================================================
# ADM CLOUD — Leer ingresos
# ======================================================================

def admcloud_get_all(resource):
    cred = base64.b64encode(
        f'{os.environ["ADMCLOUD_USER"]}:{os.environ["ADMCLOUD_PASSWORD"]}'.encode()
    ).decode()
    fijos = {
        "company": os.environ["ADMCLOUD_COMPANY"],
        "role": os.environ["ADMCLOUD_ROLE"],
        "appid": os.environ["ADMCLOUD_APPID"],
    }
    base = "https://api.admcloud.net/api"
    all_items = []
    skip = 0
    while True:
        params = {**fijos, "skip": skip}
        url = f"{base}/{resource}?{urllib.parse.urlencode(params)}"
        req = urllib.request.Request(
            url,
            headers={"Authorization": f"Basic {cred}", "Accept": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=60) as r:
            raw = json.loads(r.read().decode("utf-8", "replace"))
        if isinstance(raw, dict) and "data" in raw:
            inner = raw["data"]
            if isinstance(inner, dict) and "Item1" in inner:
                items = inner["Item1"]
            elif isinstance(inner, list):
                items = inner
            else:
                items = [inner]
        elif isinstance(raw, list):
            items = raw
        else:
            items = [raw]
        all_items.extend(items)
        if len(items) < 50:
            break
        skip += 50
        if skip > 5000:
            break
    return all_items


def load_adm_txs(start_date, end_date):
    adm = []
    for c in admcloud_get_all("CashInvoices"):
        d = (c.get("DocDate") or "")[:10]
        if start_date <= d <= end_date:
            adm.append({
                "fecha": d, "doc_id": c.get("DocID", ""),
                "tipo": "Venta de Contado",
                "descripcion": c.get("RelationshipName", c.get("CustomerName", "")),
                "monto": c.get("TotalAmount", 0) or 0,
                "referencia": c.get("Reference", "") or "",
            })
    for c in admcloud_get_all("CashReceipts"):
        d = (c.get("DocDate") or "")[:10]
        if start_date <= d <= end_date:
            adm.append({
                "fecha": d, "doc_id": c.get("DocID", ""),
                "tipo": "Recibo Cuentas por Cobrar",
                "descripcion": c.get("RelationshipName", c.get("CommercialName", "")),
                "monto": c.get("TotalAmount", 0) or 0,
                "referencia": c.get("Reference", "") or "",
            })
    for t in admcloud_get_all("BankBankTransfers"):
        d = (t.get("DocDate") or "")[:10]
        if start_date <= d <= end_date:
            adm.append({
                "fecha": d, "doc_id": t.get("DocID", ""),
                "tipo": "Transferencia entre Cuentas",
                "descripcion": f"Transf. de {t.get('FromCashAccountName', '')}",
                "monto": t.get("TotalAmount", 0) or 0,
                "referencia": t.get("Reference", "") or "",
            })
    return adm


# ======================================================================
# MOTOR DE CRUCE (8 rondas)
# ======================================================================

def match_rounds(bank, adm):
    used_bank = set()
    used_adm = set()
    matched = []

    def run_round(amt_check, max_days, require_name=False, use_original=False):
        for bi, bt in enumerate(bank):
            if bi in used_bank:
                continue
            if bt["is_credit_card"] and not use_original:
                continue
            compare_amt = bt["monto_original"] if bt["is_credit_card"] else bt["monto"]
            target = abs(compare_amt)
            best = None; best_dd = 999; best_ad = 999
            for ai, at in enumerate(adm):
                if ai in used_adm:
                    continue
                amt_diff = abs(target - abs(at["monto"]))
                if not amt_check(amt_diff, target):
                    continue
                dd = date_diff_days(bt["fecha"], at["fecha"])
                if dd > max_days:
                    continue
                if require_name and not bt["is_credit_card"]:
                    if not names_match(bt["descripcion"], at.get("descripcion", "")):
                        continue
                if dd < best_dd or (dd == best_dd and amt_diff < best_ad):
                    best_dd = dd; best_ad = amt_diff; best = ai
            if best is not None:
                used_bank.add(bi); used_adm.add(best)
                matched.append({"bank": bt, "adm": adm[best]})

    # Rondas normales (5)
    run_round(lambda d, t: d < MONTO_EXACTO_DIFF, FECHA_MAX_EXACTO, require_name=True)
    run_round(lambda d, t: d < MONTO_EXACTO_DIFF, FECHA_MAX_CERCANO)
    run_round(lambda d, t: t > 0 and d / t < MONTO_CERCANO_PCT, FECHA_MAX_CERCANO, require_name=True)
    run_round(lambda d, t: t > 0 and d / t < MONTO_CERCANO_PCT, FECHA_MAX_FUZZY)
    run_round(lambda d, t: t > 0 and d / t < MONTO_FUZZY_PCT, FECHA_MAX_FUZZY, require_name=True)
    # Rondas tarjetas (3)
    run_round(lambda d, t: d < MONTO_EXACTO_DIFF * 2, FECHA_MAX_EXACTO, use_original=True)
    run_round(lambda d, t: t > 0 and d / t < MONTO_CERCANO_PCT, FECHA_MAX_CERCANO, use_original=True)
    run_round(lambda d, t: t > 0 and d / t < MONTO_FUZZY_PCT, FECHA_MAX_FUZZY, use_original=True)

    unmatched_bank = [bank[i] for i in range(len(bank)) if i not in used_bank]
    unmatched_adm = [adm[i] for i in range(len(adm)) if i not in used_adm]
    return matched, unmatched_bank, unmatched_adm


# ======================================================================
# REPORTE — Excel
# ======================================================================

def generate_excel(matched, unmatched_bank, unmatched_adm, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl no instalado. Ejecutar: pip install openpyxl")
        return

    wb = Workbook()
    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    yfill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ofill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    gfill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    ccfill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    tf = Font(name="Calibri", bold=True, size=14, color="2F5496")
    sf = Font(name="Calibri", bold=True, size=12, color="2F5496")
    bd = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # HOJA 1: RESUMEN
    ws = wb.active; ws.title = "Resumen"
    ws["A1"] = "Entradas de Dinero: Banco vs ADM Cloud"; ws["A1"].font = tf; ws.merge_cells("A1:E1")
    ws["A2"] = "Blackbox SRL | ADM: Contado + CxC + Transferencias | Tarjeta: comision 5.395% revertida"
    ws["A2"].font = Font(italic=True, size=10); ws.merge_cells("A2:E2")
    ws["A3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].font = Font(italic=True, size=9, color="808080"); ws.merge_cells("A3:E3")

    r = 5; ws.cell(row=r, column=1, value="RESUMEN").font = sf; r += 1
    for col, h in enumerate(["Metrica", "Cantidad", "Monto DOP", "", ""], 1):
        c = ws.cell(row=r, column=col, value=h)
        if h: c.font = hf; c.fill = hfill; c.border = bd
    r += 1
    total_bank = sum(m["bank"]["monto"] for m in matched) + sum(t["monto"] for t in unmatched_bank)
    total_adm = sum(m["adm"]["monto"] for m in matched) + sum(t["monto"] for t in unmatched_adm)
    total_matched = sum(m["bank"]["monto"] for m in matched)
    stats = [
        ("Entradas en el banco", len(matched) + len(unmatched_bank), f"{total_bank:,.2f}"),
        ("Entradas en ADM Cloud", len(matched) + len(unmatched_adm), f"{total_adm:,.2f}"),
        ("", "", ""),
        ("CONCILIADOS (banco = ADM)", len(matched), f"{total_matched:,.2f}"),
        ("EN BANCO, sin registro en ADM", len(unmatched_bank), f"{sum(t['monto'] for t in unmatched_bank):,.2f}"),
        ("EN ADM, sin entrada en banco", len(unmatched_adm), f"{sum(t['monto'] for t in unmatched_adm):,.2f}"),
    ]
    for label, cnt, monto in stats:
        if label:
            ws.cell(row=r, column=1, value=label).border = bd
            ws.cell(row=r, column=2, value=cnt).border = bd
            ws.cell(row=r, column=3, value=monto).border = bd
            if "CONCILIADOS" in label:
                for col in range(1, 4): ws.cell(row=r, column=col).fill = gfill
            elif "BANCO" in label:
                for col in range(1, 4): ws.cell(row=r, column=col).fill = yfill
            elif "ADM" in label and "ADM Cloud" not in label:
                for col in range(1, 4): ws.cell(row=r, column=col).fill = ofill
        r += 1
    ws.column_dimensions["A"].width = 38; ws.column_dimensions["B"].width = 12; ws.column_dimensions["C"].width = 18

    # HOJA 2: ENTRADAS DEL BANCO
    ws2 = wb.create_sheet("Entradas Banco")
    headers = ["#", "Fecha", "Cuenta", "Descripcion", "Monto", "Monto Orig.", "Nro.Ref.", "Tarjeta?", "Estado", "Match ADM"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h); c.font = hf; c.fill = hfill; c.border = bd
    all_bank = [(m["bank"], True, m["adm"]) for m in matched] + [(t, False, None) for t in unmatched_bank]
    all_bank.sort(key=lambda x: x[0]["fecha"], reverse=True)
    for i, (tx, is_m, adm) in enumerate(all_bank, 2):
        is_cc = tx.get("is_credit_card", False)
        vals = [i - 1, tx["fecha"], CUENTA_NOMBRES.get(tx.get("cuenta", ""), tx.get("cuenta", "")),
                tx["descripcion"], tx["monto"],
                round(tx.get("monto_original", tx["monto"]), 2) if is_cc else "",
                tx.get("nro_referencia", ""), "Si" if is_cc else "",
                "Conciliado" if is_m else "NO conciliado",
                f'{adm["doc_id"]} - {adm["tipo"]}' if adm else ""]
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row=i, column=col, value=v); c.border = bd
            if col in (5, 6) and isinstance(v, (int, float)): c.number_format = "#,##0.00"
            c.fill = ccfill if is_cc else (gfill if is_m else yfill)
    ws2.freeze_panes = "A2"
    for col, w in enumerate([5, 12, 15, 45, 14, 14, 14, 10, 14, 35], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # HOJA 3: ENTRADAS ADM
    ws3 = wb.create_sheet("Entradas ADM")
    headers = ["#", "Fecha", "Doc ADM", "Tipo", "Cliente", "Monto", "Referencia", "Estado", "Match Banco"]
    for col, h in enumerate(headers, 1):
        c = ws3.cell(row=1, column=col, value=h); c.font = hf; c.fill = hfill; c.border = bd
    all_adm = [(m["adm"], True, m["bank"]) for m in matched] + [(t, False, None) for t in unmatched_adm]
    all_adm.sort(key=lambda x: x[0]["fecha"], reverse=True)
    for i, (tx, is_m, bk) in enumerate(all_adm, 2):
        vals = [i - 1, tx["fecha"], tx.get("doc_id", ""), tx.get("tipo", ""),
                tx.get("descripcion", ""), tx["monto"], tx.get("referencia", ""),
                "Conciliado" if is_m else "NO conciliado",
                bk["descripcion"][:50] if bk else ""]
        for col, v in enumerate(vals, 1):
            c = ws3.cell(row=i, column=col, value=v); c.border = bd
            if col == 6: c.number_format = "#,##0.00"
            c.fill = gfill if is_m else ofill
    ws3.freeze_panes = "A2"
    for col, w in enumerate([5, 12, 14, 25, 30, 14, 14, 14, 50], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)


# ======================================================================
# MAIN
# ======================================================================

def end_of_month(year, month):
    if month == 12:
        return date(year, 12, 31)
    return date(year, month + 1, 1) - timedelta(days=1)


def main():
    if len(sys.argv) >= 3:
        y1, m1 = int(sys.argv[1][:4]), int(sys.argv[1][5:7])
        y2, m2 = int(sys.argv[2][:4]), int(sys.argv[2][5:7])
        start_date = date(y1, m1, 1).isoformat()
        end_date = end_of_month(y2, m2).isoformat()
        label = f"{sys.argv[1]}_a_{sys.argv[2]}"
    elif len(sys.argv) == 2:
        y, m = int(sys.argv[1][:4]), int(sys.argv[1][5:7])
        start_date = date(y, m, 1).isoformat()
        end_date = end_of_month(y, m).isoformat()
        label = sys.argv[1]
    else:
        today = date.today()
        start_date = date(today.year, today.month, 1).isoformat()
        end_date = end_of_month(today.year, today.month).isoformat()
        label = f"{today.year}-{today.month:02d}"

    print(f"Período: {start_date} a {end_date}\n")

    print("Descargando banco (Supabase)...", end=" ", flush=True)
    bank = load_bank_txs(start_date, end_date)
    print(f"{len(bank)} entradas")

    print("Descargando ADM Cloud... (puede tardar)", end=" ", flush=True)
    adm = load_adm_txs(start_date, end_date)
    print(f"{len(adm)} entradas")

    cc_count = sum(1 for t in bank if t["is_credit_card"])
    print(f"  ({cc_count} tarjeta, {len(bank) - cc_count} normal)\n")

    print("Cruzando...", end=" ", flush=True)
    matched, unmatched_bank, unmatched_adm = match_rounds(bank, adm)
    print("listo\n")

    total_bank = sum(m["bank"]["monto"] for m in matched) + sum(t["monto"] for t in unmatched_bank)
    total_adm = sum(m["adm"]["monto"] for m in matched) + sum(t["monto"] for t in unmatched_adm)
    total_matched = sum(m["bank"]["monto"] for m in matched)

    print("=" * 60)
    print(f"  ENTRADAS BANCO:           {len(matched)+len(unmatched_bank):>4}  {total_bank:>14,.2f}")
    print(f"  ENTRADAS ADM:             {len(matched)+len(unmatched_adm):>4}  {total_adm:>14,.2f}")
    print(f"  CONCILIADOS:              {len(matched):>4}  {total_matched:>14,.2f}")
    print(f"  EN BANCO SIN ADM:         {len(unmatched_bank):>4}  {sum(t['monto'] for t in unmatched_bank):>14,.2f}")
    print(f"  EN ADM SIN BANCO:         {len(unmatched_adm):>4}  {sum(t['monto'] for t in unmatched_adm):>14,.2f}")
    print("=" * 60 + "\n")

    excel_path = f"/opt/data/Conciliacion_Entradas_{label}.xlsx"
    print(f"Generando Excel...", end=" ", flush=True)
    generate_excel(matched, unmatched_bank, unmatched_adm, excel_path)
    print(f"{excel_path}")

    json_path = f"/opt/data/conciliacion_{label}.json"
    json.dump(
        {"matched": matched, "unmatched_bank": unmatched_bank, "unmatched_adm": unmatched_adm},
        open(json_path, "w"), indent=2, default=str, ensure_ascii=False,
    )
    print(f"JSON: {json_path}")


if __name__ == "__main__":
    main()
